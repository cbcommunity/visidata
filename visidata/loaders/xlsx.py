
from visidata import *

class open_xlsx(Sheet):
    'Load XLSX file (in Excel Open XML format).'
    commands = [
        Command(ENTER, 'vd.push(sheet.getSheet(cursorRow))', 'load the entire table into memory')
    ]
    def __init__(self, path):
        super().__init__(path.name, source=path)
        self.workbook = None

    @async
    def reload(self):
        import openpyxl
        self.columns = [Column('name')]
        self.workbook = openpyxl.load_workbook(self.source.resolve(), data_only=True, read_only=True)
        self.rows = list(self.workbook.sheetnames)

    def getSheet(self, sheetname):
        worksheet = self.workbook.get_sheet_by_name(sheetname)
        return xlsxSheet(joinSheetnames(self.name, sheetname), source=worksheet)

class xlsxSheet(Sheet):
    @async
    def reload(self):
        worksheet = self.source

        self.columns = []
        self.rows = []
        for row in Progress(worksheet.iter_rows(), worksheet.max_row or 0):
            L = list(cell.value for cell in row)
            for i in range(len(self.columns), len(L)):  # no-op if already done
                self.addColumn(ColumnItem(None, i, width=8))
            self.addRow(L)

class open_xls(Sheet):
    'Load XLS file (in Excel format).'
    commands = [
        Command(ENTER, 'vd.push(sheet.getSheet(cursorRow))', 'load the entire table into memory')
    ]
    def __init__(self, path):
        super().__init__(path.name, source=path)
        self.workbook = None

    @async
    def reload(self):
        import xlrd
        self.columns = [Column('name')]
        self.workbook = xlrd.open_workbook(self.source.resolve())
        self.rows = list(self.workbook.sheet_names())

    def getSheet(self, sheetname):
        worksheet = self.workbook.sheet_by_name(sheetname)
        return xlsSheet(joinSheetnames(self.name, sheetname), source=worksheet)


class xlsSheet(Sheet):
    @async
    def reload(self):
        worksheet = self.source
        self.columns = []
        for i in range(worksheet.ncols):
            self.addColumn(ColumnItem(None, i, width=8))

        for rownum in Progress(range(worksheet.nrows)):
            self.addRow(list(worksheet.cell(rownum, colnum).value for colnum in range(worksheet.ncols)))
